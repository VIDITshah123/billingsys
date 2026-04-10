from flask import Flask, render_template, request, redirect, url_for, flash, send_file, make_response, jsonify, session
from flask_sqlalchemy import SQLAlchemy
from flask_wtf import FlaskForm
from wtforms import StringField, TextAreaField, SelectField, DecimalField, DateField, SelectMultipleField, widgets
from wtforms.validators import DataRequired, Length, Regexp, NumberRange
from datetime import datetime, date
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER
from reportlab.lib.units import inch, mm
from reportlab.lib.colors import black, blue, red
from io import BytesIO, StringIO
import os
import json
from functools import wraps

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-change-in-production'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///billingsys.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

COMPANY_INFO = {
    'name': 'VIDHIM ENTERPRISES',
    'address': 'FIRST FLOOR, 105, BHAURAO UDYOG NAGAR, KHARIGAON , ABOVE S K STEEL, BHAYANDER (E)-401105',
    'phone': '+91 9892352600',
    'email': 'vidhimenterprises@gmail.com',
    'gst_no': '27AXVPS9856J1Z4'
}

USERS_FILE = 'users.json'

def load_users():
    if os.path.exists(USERS_FILE):
        with open(USERS_FILE, 'r') as f:
            return json.load(f)
    else:
        default_users = [{"userid": "admin", "password": "admin123"}]
        with open(USERS_FILE, 'w') as f:
            json.dump(default_users, f, indent=4)
        return default_users

def validate_user(userid, password):
    users = load_users()
    for user in users:
        if user.get('userid') == userid and user.get('password') == password:
            return True
    return False

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

class Customer(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    address = db.Column(db.Text, nullable=False)
    gst_number = db.Column(db.String(15), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    invoices = db.relationship('Invoice', backref='customer', lazy=True)

    def __repr__(self):
        return f'<Customer {self.name}>'

class Product(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    hsn_code = db.Column(db.String(8), nullable=False)
    unit = db.Column(db.String(10), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    def __repr__(self):
        return f'<Product {self.name}>'

class Invoice(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    invoice_number = db.Column(db.String(50), unique=True, nullable=False)
    invoice_date = db.Column(db.Date, nullable=False)
    customer_id = db.Column(db.Integer, db.ForeignKey('customer.id'), nullable=False)
    tax_type = db.Column(db.String(20), nullable=False, default='cgst_sgst')
    subtotal = db.Column(db.Numeric(10, 2), default=0)
    cgst = db.Column(db.Numeric(10, 2), default=0)
    sgst = db.Column(db.Numeric(10, 2), default=0)
    igst = db.Column(db.Numeric(10, 2), default=0)
    total = db.Column(db.Numeric(10, 2), default=0)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    items = db.relationship('InvoiceItem', backref='invoice', lazy=True, cascade='all, delete-orphan')

    def calculate_taxes(self):
        from decimal import Decimal
        self.subtotal = sum(Decimal(str(item.quantity)) * Decimal(str(item.rate)) for item in self.items)
        
        if self.tax_type == 'cgst_sgst':
            self.cgst = self.subtotal * Decimal('0.025')
            self.sgst = self.subtotal * Decimal('0.025')
            self.igst = Decimal('0')
        else:
            self.igst = self.subtotal * Decimal('0.05')
            self.cgst = Decimal('0')
            self.sgst = Decimal('0')
        
        total_with_taxes = self.subtotal + self.cgst + self.sgst + self.igst
        self.total = int(total_with_taxes.quantize(Decimal('1')))

    def __repr__(self):
        return f'<Invoice {self.invoice_number}>'

class InvoiceItem(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    invoice_id = db.Column(db.Integer, db.ForeignKey('invoice.id'), nullable=False)
    product_id = db.Column(db.Integer, db.ForeignKey('product.id'), nullable=False)
    quantity = db.Column(db.Numeric(10, 2), nullable=False)
    rate = db.Column(db.Numeric(10, 2), nullable=False)
    total = db.Column(db.Numeric(10, 2), nullable=False)
    
    product = db.relationship('Product', backref='invoice_items')

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        if self.quantity and self.rate:
            self.total = self.quantity * self.rate

class MultiCheckboxField(SelectMultipleField):
    widget = widgets.ListWidget(prefix_label=False)
    option_widget = widgets.CheckboxInput()

class CustomerForm(FlaskForm):
    name = StringField('Name', validators=[DataRequired(), Length(min=2, max=200)])
    address = TextAreaField('Address', validators=[DataRequired()])
    gst_number = StringField('GST Number', validators=[
        DataRequired(), 
        Regexp(r'^[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z]{1}[1-9A-Z]{1}Z[0-9A-Z]{1}$', message='Invalid GST number format')
    ])

class ProductForm(FlaskForm):
    name = StringField('Name', validators=[DataRequired(), Length(min=2, max=200)])
    hsn_code = StringField('HSN Code', validators=[DataRequired(), Length(min=4, max=8)])
    unit = SelectField('Unit', choices=[('kgs', 'Kilograms'), ('units', 'Units')], validators=[DataRequired()])

class InvoiceItemForm(FlaskForm):
    product = SelectField('Product', coerce=int, validators=[DataRequired()])
    quantity = DecimalField('Quantity', validators=[DataRequired(), NumberRange(min=0)], places=2)
    rate = DecimalField('Rate', validators=[DataRequired(), NumberRange(min=0)], places=2)

class InvoiceForm(FlaskForm):
    invoice_number = StringField('Invoice Number', validators=[DataRequired()])
    invoice_date = DateField('Invoice Date', validators=[DataRequired()], default=date.today)
    customer = SelectField('Customer', coerce=int, validators=[DataRequired()])
    tax_type = SelectField('Tax Type', choices=[
        ('cgst_sgst', 'CGST + SGST'), 
        ('igst', 'IGST')
    ], validators=[DataRequired()])

class ReportForm(FlaskForm):
    start_date = DateField('Start Date')
    end_date = DateField('End Date')
    customer = SelectField('Customer', coerce=int)
    include_fields = MultiCheckboxField('Include Fields', choices=[
        ('invoice_number', 'Invoice Number'),
        ('invoice_date', 'Invoice Date'),
        ('customer', 'Customer'),
        ('subtotal', 'Subtotal'),
        ('cgst', 'CGST'),
        ('sgst', 'SGST'),
        ('igst', 'IGST'),
        ('total', 'Total'),
    ])

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        userid = request.form.get('userid')
        password = request.form.get('password')
        
        if validate_user(userid, password):
            session['logged_in'] = True
            session['userid'] = userid
            return redirect(url_for('home'))
        else:
            flash('Invalid credentials. Please try again.', 'danger')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.pop('logged_in', None)
    session.pop('userid', None)
    return redirect(url_for('login'))

@app.route('/')
@login_required
def home():
    return render_template('home.html', company_info=COMPANY_INFO)

@app.route('/customers')
@login_required
def customer_list():
    search = request.args.get('search', '')
    customers = Customer.query
    
    if search:
        customers = customers.filter(
            Customer.name.contains(search) | 
            Customer.gst_number.contains(search)
        )
    
    customers = customers.order_by(Customer.created_at.desc()).all()
    return render_template('customer_list.html', customers=customers, search=search)

@app.route('/customers/inline-add', methods=['POST'])
def customer_inline_add():
    try:
        name = request.form.get('name')
        address = request.form.get('address')
        gst_number = request.form.get('gst_number', '').upper()
        
        if not name or not address or not gst_number:
            return jsonify({'success': False, 'message': 'All fields are required'})
        
        import re
        gst_pattern = r'^[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z]{1}[1-9A-Z]{1}Z[0-9A-Z]{1}$'
        if not re.match(gst_pattern, gst_number):
            return jsonify({'success': False, 'message': 'Invalid GST number format'})
        
        customer = Customer(name=name, address=address, gst_number=gst_number)
        db.session.add(customer)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Customer added successfully!'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

@app.route('/customers/create', methods=['GET', 'POST'])
@login_required
def customer_create():
    form = CustomerForm()
    if form.validate_on_submit():
        customer = Customer(
            name=form.name.data,
            address=form.address.data,
            gst_number=form.gst_number.data.upper()
        )
        db.session.add(customer)
        db.session.commit()
        flash('Customer created successfully!', 'success')
        return redirect(url_for('customer_list'))
    return render_template('customer_form.html', form=form)

@app.route('/customers/<int:id>/update', methods=['GET', 'POST'])
@login_required
def customer_update(id):
    customer = Customer.query.get_or_404(id)
    form = CustomerForm(obj=customer)
    if form.validate_on_submit():
        customer.name = form.name.data
        customer.address = form.address.data
        customer.gst_number = form.gst_number.data.upper()
        db.session.commit()
        flash('Customer updated successfully!', 'success')
        return redirect(url_for('customer_list'))
    return render_template('customer_form.html', form=form, customer=customer)

@app.route('/customers/<int:id>/delete', methods=['POST'])
@login_required
def customer_delete(id):
    customer = Customer.query.get_or_404(id)
    db.session.delete(customer)
    db.session.commit()
    flash('Customer deleted successfully!', 'success')
    return redirect(url_for('customer_list'))

@app.route('/products')
@login_required
def product_list():
    search = request.args.get('search', '')
    products = Product.query
    
    if search:
        products = products.filter(
            Product.name.contains(search) | 
            Product.hsn_code.contains(search)
        )
    
    products = products.order_by(Product.created_at.desc()).all()
    return render_template('product_list.html', products=products, search=search)

@app.route('/products/inline-add', methods=['POST'])
def product_inline_add():
    try:
        name = request.form.get('name')
        hsn_code = request.form.get('hsn_code')
        unit = request.form.get('unit')
        
        if not name or not hsn_code or not unit:
            return jsonify({'success': False, 'message': 'All fields are required'})
        
        if not hsn_code.isdigit() or len(hsn_code) < 4 or len(hsn_code) > 8:
            return jsonify({'success': False, 'message': 'HSN code should be 4-8 digits'})
        
        product = Product(name=name, hsn_code=hsn_code, unit=unit)
        db.session.add(product)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Product added successfully!'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

@app.route('/products/create', methods=['GET', 'POST'])
@login_required
def product_create():
    form = ProductForm()
    if form.validate_on_submit():
        product = Product(
            name=form.name.data,
            hsn_code=form.hsn_code.data,
            unit=form.unit.data
        )
        db.session.add(product)
        db.session.commit()
        flash('Product created successfully!', 'success')
        return redirect(url_for('product_list'))
    return render_template('product_form.html', form=form)

@app.route('/products/<int:id>/update', methods=['GET', 'POST'])
@login_required
def product_update(id):
    product = Product.query.get_or_404(id)
    form = ProductForm(obj=product)
    if form.validate_on_submit():
        product.name = form.name.data
        product.hsn_code = form.hsn_code.data
        product.unit = form.unit.data
        db.session.commit()
        flash('Product updated successfully!', 'success')
        return redirect(url_for('product_list'))
    return render_template('product_form.html', form=form, product=product)

@app.route('/products/<int:id>/delete', methods=['POST'])
@login_required
def product_delete(id):
    product = Product.query.get_or_404(id)
    db.session.delete(product)
    db.session.commit()
    flash('Product deleted successfully!', 'success')
    return redirect(url_for('product_list'))

@app.route('/invoices')
@login_required
def invoice_list():
    search = request.args.get('search', '')
    invoices = Invoice.query
    
    if search:
        invoices = invoices.filter(
            Invoice.invoice_number.contains(search) | 
            Customer.name.contains(search)
        ).join(Customer)
    
    invoices = invoices.order_by(Invoice.created_at.desc()).all()
    return render_template('invoice_list.html', invoices=invoices, search=search)

@app.route('/invoices/create', methods=['GET', 'POST'])
@login_required
def invoice_create():
    form = InvoiceForm()
    form.customer.choices = [(c.id, c.name) for c in Customer.query.order_by(Customer.name).all()]
    products = Product.query.order_by(Product.name).all()
    
    if request.method == 'POST':
        invoice = Invoice(
            invoice_number=form.invoice_number.data,
            invoice_date=form.invoice_date.data,
            customer_id=form.customer.data,
            tax_type=form.tax_type.data
        )
        db.session.add(invoice)
        db.session.commit()
        flash('Invoice created! Please add items to the invoice.', 'success')
        return redirect(url_for('invoice_detail', id=invoice.id))
    
    return render_template('invoice_form.html', form=form, products=products)

@app.route('/invoices/create-with-items', methods=['POST'])
@login_required
def invoice_create_with_items():
    try:
        invoice_number = request.form.get('invoice_number')
        invoice_date_str = request.form.get('invoice_date')
        customer_id = int(request.form.get('customer'))
        tax_type = request.form.get('tax_type')
        items_json = request.form.get('items')
        
        if not invoice_number or not invoice_date_str or not customer_id or not items_json:
            return jsonify({'success': False, 'message': 'All fields are required'})
        
        invoice_date = datetime.strptime(invoice_date_str, '%Y-%m-%d').date()
        items_data = json.loads(items_json)
        
        if len(items_data) == 0:
            return jsonify({'success': False, 'message': 'At least one item is required'})
        
        invoice = Invoice(
            invoice_number=invoice_number,
            invoice_date=invoice_date,
            customer_id=customer_id,
            tax_type=tax_type
        )
        db.session.add(invoice)
        db.session.flush()
        
        for item in items_data:
            invoice_item = InvoiceItem(
                invoice_id=invoice.id,
                product_id=item['product_id'],
                quantity=item['quantity'],
                rate=item['rate']
            )
            db.session.add(invoice_item)
        
        invoice.calculate_taxes()
        db.session.commit()
        
        return jsonify({'success': True, 'redirect_url': url_for('invoice_detail', id=invoice.id)})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

@app.route('/invoices/<int:id>')
@login_required
def invoice_detail(id):
    invoice = Invoice.query.get_or_404(id)
    products = Product.query.order_by(Product.name).all()
    return render_template('invoice_detail.html', invoice=invoice, products=products, company_info=COMPANY_INFO)

@app.route('/invoices/<int:id>/add_item', methods=['POST'])
@login_required
def add_invoice_item(id):
    invoice = Invoice.query.get_or_404(id)
    product_id = request.form.get('product_id')
    quantity = float(request.form.get('quantity', 0))
    rate = float(request.form.get('rate', 0))
    
    if product_id and quantity > 0 and rate > 0:
        product = Product.query.get(product_id)
        if product:
            item = InvoiceItem(
                invoice_id=invoice.id,
                product_id=product.id,
                quantity=quantity,
                rate=rate
            )
            db.session.add(item)
            invoice.calculate_taxes()
            db.session.commit()
            flash('Item added successfully!', 'success')
    
    return redirect(url_for('invoice_detail', id=id))

@app.route('/invoices/<int:id>/delete_item/<int:item_id>', methods=['POST'])
@login_required
def delete_invoice_item(id, item_id):
    try:
        invoice = Invoice.query.get_or_404(id)
        item = InvoiceItem.query.get_or_404(item_id)
        
        if item.invoice_id == invoice.id:
            db.session.delete(item)
            invoice.calculate_taxes()
            db.session.commit()
            flash('Item deleted successfully!', 'success')
        else:
            flash('Item not found in this invoice!', 'error')
    
    except Exception as e:
        flash(f'Error deleting item: {str(e)}', 'error')
    
    return redirect(url_for('invoice_detail', id=id))

@app.route('/invoices/<int:id>/delete', methods=['POST'])
@login_required
def invoice_delete(id):
    invoice = Invoice.query.get_or_404(id)
    db.session.delete(invoice)
    db.session.commit()
    flash('Invoice deleted successfully!', 'success')
    return redirect(url_for('invoice_list'))

@app.route('/invoices/<int:id>/pdf')
@login_required
def generate_invoice_pdf(id):
    try:
        invoice = Invoice.query.get_or_404(id)
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=A4,
            rightMargin=40,
            leftMargin=40,
            topMargin=40,
            bottomMargin=40
        )
        story = []
        styles = getSampleStyleSheet()
        width, height = A4
        
        primary_color = colors.HexColor('#1a365d')
        accent_color = colors.HexColor('#2b6cb0')
        light_bg = colors.HexColor('#f7fafc')
        border_color = colors.HexColor('#e2e8f0')
        
        header_data = [[
            Paragraph(
                f"<font size='18' color='#1a365d'><b>{COMPANY_INFO['name']}</b></font>",
                styles['Normal']
            ),
            Paragraph(
                "<font size='28' color='#1a365d'><b>INVOICE</b></font>",
                ParagraphStyle('InvoiceTitle', alignment=2)
            )
        ]]
        
        header_table = Table(header_data, colWidths=[3.5*inch, 3.5*inch])
        header_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ]))
        story.append(header_table)
        story.append(Spacer(1, 0.15*inch))
        
        line_table = Table([['']], colWidths=[7*inch])
        line_table.setStyle(TableStyle([
            ('LINEABOVE', (0, 0), (-1, 0), 3, accent_color),
        ]))
        story.append(line_table)
        story.append(Spacer(1, 0.25*inch))
        
        info_style = ParagraphStyle('InfoStyle', parent=styles['Normal'], fontSize=9, leading=14)
        label_style = ParagraphStyle('LabelStyle', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor('#718096'), leading=12)
        
        invoice_info = [
            [Paragraph("<b>Invoice Number</b>", label_style), Paragraph("<b>Invoice Date</b>", label_style)],
            [Paragraph(f"<b>{invoice.invoice_number}</b>", info_style), Paragraph(invoice.invoice_date.strftime('%B %d, %Y'), info_style)]
        ]
        
        info_table = Table(invoice_info, colWidths=[3.5*inch, 3.5*inch])
        info_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 2),
            ('TOPPADDING', (0, 1), (-1, 1), 2),
        ]))
        story.append(info_table)
        story.append(Spacer(1, 0.3*inch))
        
        section_title = ParagraphStyle('SectionTitle', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor('#718096'), spaceBefore=0, spaceAfter=6)
        address_style = ParagraphStyle('AddressStyle', parent=styles['Normal'], fontSize=9, leading=14)
        
        from_to_data = [[
            [Paragraph("FROM", section_title), Paragraph(f"<b>{COMPANY_INFO['name']}</b>", address_style), Paragraph(COMPANY_INFO['address'], address_style), Paragraph(f"Phone: {COMPANY_INFO['phone']}", address_style), Paragraph(f"Email: {COMPANY_INFO['email']}", address_style), Paragraph(f"<b>GSTIN:</b> {COMPANY_INFO['gst_no']}", address_style)],
            [Paragraph("BILL TO", section_title), Paragraph(f"<b>{invoice.customer.name}</b>", address_style), Paragraph(invoice.customer.address, address_style), Paragraph(f"<b>GSTIN:</b> {invoice.customer.gst_number}", address_style)]
        ]]
        
        from_content = Table([[p] for p in from_to_data[0][0]], colWidths=[3.2*inch])
        to_content = Table([[p] for p in from_to_data[0][1]], colWidths=[3.2*inch])
        from_to_table = Table([[from_content, to_content]], colWidths=[3.5*inch, 3.5*inch])
        from_to_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (0, 0), 0),
            ('RIGHTPADDING', (1, 0), (1, 0), 0),
        ]))
        story.append(from_to_table)
        story.append(Spacer(1, 0.35*inch))
        
        header_style = ParagraphStyle('TableHeader', parent=styles['Normal'], fontSize=8, textColor=colors.white, alignment=1)
        cell_style = ParagraphStyle('CellStyle', parent=styles['Normal'], fontSize=9, leading=12)
        cell_style_right = ParagraphStyle('CellStyleRight', parent=styles['Normal'], fontSize=9, leading=12, alignment=2)
        cell_style_center = ParagraphStyle('CellStyleCenter', parent=styles['Normal'], fontSize=9, leading=12, alignment=1)
        
        items_data = [[Paragraph("<b>#</b>", header_style), Paragraph("<b>DESCRIPTION</b>", header_style), Paragraph("<b>HSN</b>", header_style), Paragraph("<b>QTY</b>", header_style), Paragraph("<b>RATE</b>", header_style), Paragraph("<b>AMOUNT</b>", header_style)]]
        
        for i, item in enumerate(invoice.items, 1):
            items_data.append([Paragraph(str(i), cell_style_center), Paragraph(item.product.name, cell_style), Paragraph(item.product.hsn_code, cell_style_center), Paragraph(f"{item.quantity:.2f} {item.product.unit}", cell_style_center), Paragraph(f"Rs.{item.rate:,.2f}", cell_style_right), Paragraph(f"Rs.{item.total:,.2f}", cell_style_right)])
        
        items_table = Table(items_data, colWidths=[0.4*inch, 2.4*inch, 0.8*inch, 0.9*inch, 1*inch, 1.1*inch])
        items_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), primary_color),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('TOPPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            *[('BACKGROUND', (0, i), (-1, i), light_bg) for i in range(2, len(items_data), 2)],
            ('LINEBELOW', (0, 0), (-1, 0), 1, primary_color),
            ('LINEBELOW', (0, 1), (-1, -2), 0.5, border_color),
            ('LINEBELOW', (0, -1), (-1, -1), 1, primary_color),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (2, 0), (3, -1), 'CENTER'),
            ('ALIGN', (4, 1), (-1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(items_table)
        story.append(Spacer(1, 0.25*inch))
        
        totals_label = ParagraphStyle('TotalsLabel', parent=styles['Normal'], fontSize=9, alignment=2)
        totals_value = ParagraphStyle('TotalsValue', parent=styles['Normal'], fontSize=9, alignment=2)
        totals_final_label = ParagraphStyle('TotalsFinalLabel', parent=styles['Normal'], fontSize=11, alignment=2, textColor=primary_color)
        totals_final_value = ParagraphStyle('TotalsFinalValue', parent=styles['Normal'], fontSize=11, alignment=2, textColor=primary_color)
        
        totals_data = [[Paragraph("Subtotal", totals_label), Paragraph(f"Rs.{invoice.subtotal:,.2f}", totals_value)]]
        
        if invoice.tax_type == 'cgst_sgst':
            totals_data.extend([[Paragraph("CGST @ 2.5%", totals_label), Paragraph(f"Rs.{invoice.cgst:,.2f}", totals_value)], [Paragraph("SGST @ 2.5%", totals_label), Paragraph(f"Rs.{invoice.sgst:,.2f}", totals_value)]])
        else:
            totals_data.append([Paragraph("IGST @ 5%", totals_label), Paragraph(f"Rs.{invoice.igst:,.2f}", totals_value)])
        
        totals_data.append([Paragraph("<b>TOTAL</b>", totals_final_label), Paragraph(f"<b>Rs. {int(invoice.total):,}</b>", totals_final_value)])
        
        totals_table = Table(totals_data, colWidths=[1.5*inch, 1.3*inch])
        totals_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -2), 6),
            ('TOPPADDING', (0, 0), (-1, -2), 6),
            ('LINEABOVE', (0, -1), (-1, -1), 1.5, primary_color),
            ('BOTTOMPADDING', (0, -1), (-1, -1), 8),
            ('TOPPADDING', (0, -1), (-1, -1), 8),
            ('BACKGROUND', (0, -1), (-1, -1), light_bg),
        ]))
        
        wrapper_data = [['', totals_table]]
        wrapper_table = Table(wrapper_data, colWidths=[4.2*inch, 2.8*inch])
        wrapper_table.setStyle(TableStyle([('ALIGN', (1, 0), (1, 0), 'RIGHT'), ('VALIGN', (0, 0), (-1, -1), 'TOP')]))
        story.append(wrapper_table)
        story.append(Spacer(1, 0.4*inch))
        
        footer_title = ParagraphStyle('FooterTitle', parent=styles['Normal'], fontSize=9, textColor=primary_color, spaceBefore=0, spaceAfter=6)
        footer_text = ParagraphStyle('FooterText', parent=styles['Normal'], fontSize=8, leading=12, textColor=colors.HexColor('#4a5568'))
        
        bank_details = Table([[Paragraph("<b>Bank Details</b>", footer_title)], [Paragraph("UNION BANK OF INDIA,BHAYANDAR EAST,JESAL PARK BRANCH", footer_text)], [Paragraph("Account No: 510101006809654", footer_text)], [Paragraph("IFSC: UBIN0904554", footer_text)]], colWidths=[3*inch])
        terms = Table([[Paragraph("<b>Terms & Conditions</b>", footer_title)], [Paragraph("1. Payment requested by crossed cheque payee A/c cheque/NEFT/RTGS only", footer_text)], [Paragraph("2. Our responsibility ceases on delivery of the goods to transport", footer_text)], [Paragraph("3. Goods supplied to order will not be accepted back", footer_text)], [Paragraph("4. Subject to Mumbai Jurisdiction", footer_text)], [Paragraph("5. Interest @24% p.a. will be charge on bill remaining unpaid after due date", footer_text)]], colWidths=[3.5*inch])
        footer_table = Table([[bank_details, terms]], colWidths=[3.5*inch, 3.5*inch])
        footer_table.setStyle(TableStyle([('ALIGN', (0, 0), (-1, -1), 'LEFT'), ('VALIGN', (0, 0), (-1, -1), 'TOP')]))
        story.append(footer_table)
        story.append(Spacer(1, 0.3*inch))
        
        signature_style = ParagraphStyle('Signature', parent=styles['Normal'], fontSize=9, alignment=2)
        sig_table = Table([['', Paragraph("_" * 30, signature_style)], ['', Paragraph(f"For <b>{COMPANY_INFO['name']}</b>", signature_style)], ['', Paragraph("Authorized Signatory", footer_text)]], colWidths=[4.5*inch, 2.5*inch])
        sig_table.setStyle(TableStyle([('ALIGN', (1, 0), (1, -1), 'RIGHT'), ('TOPPADDING', (1, 0), (1, 0), 20)]))
        story.append(sig_table)
        story.append(Spacer(1, 0.3*inch))
        thank_you_style = ParagraphStyle('ThankYou', parent=styles['Normal'], fontSize=10, alignment=1, textColor=colors.HexColor('#718096'))
        story.append(Paragraph("Thank you for your business!", thank_you_style))
        
        doc.build(story)
        pdf = buffer.getvalue()
        buffer.close()
        
        response = make_response(pdf)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename="Invoice_{invoice.invoice_number}.pdf"'
        response.headers['Content-Length'] = len(pdf)
        return response
        
    except Exception as e:
        flash(f'Error generating PDF: {str(e)}', 'error')
        return redirect(url_for('invoice_detail', id=id))

@app.route('/invoices/<int:id>/save')
@login_required
def save_invoice(id):
    try:
        invoice = Invoice.query.get_or_404(id)
        
        summary = f"""
VIDHIM ENTERPRISES - INVOICE SUMMARY
=====================================

Invoice Number: {invoice.invoice_number}
Invoice Date: {invoice.invoice_date.strftime('%d-%m-%Y')}
Customer: {invoice.customer.name}
Customer Address: {invoice.customer.address}
Customer GST: {invoice.customer.gst_number}

Tax Type: {invoice.tax_type.replace('_', ' ').upper()}

ITEMS:
-------
"""
        
        for i, item in enumerate(invoice.items, 1):
            summary += f"""
{i}. {item.product.name}
   HSN Code: {item.product.hsn_code}
   Quantity: {item.quantity:.2f} {item.product.unit}
   Rate: Rs. {item.rate:.2f}
   Total: Rs. {item.total:.2f}
"""
        
        summary += f"""
TAX CALCULATION:
----------------
Subtotal: Rs. {invoice.subtotal:.2f}
CGST (2.5%): Rs. {invoice.cgst:.2f}
SGST (2.5%): Rs. {invoice.sgst:.2f}
IGST (5%): Rs. {invoice.igst:.2f}
TOTAL: Rs. {invoice.total:.2f}

Generated on: {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}
"""
        
        response = make_response(summary)
        response.headers['Content-Type'] = 'text/plain'
        response.headers['Content-Disposition'] = f'attachment; filename="Invoice_{invoice.invoice_number}_Summary.txt"'
        return response
        
    except Exception as e:
        flash(f'Error saving invoice: {str(e)}', 'error')
        return redirect(url_for('invoice_detail', id=id))

@app.route('/reports', methods=['GET', 'POST'])
@login_required
def reports():
    form = ReportForm()
    form.customer.choices = [(0, 'All Customers')] + [(c.id, c.name) for c in Customer.query.order_by(Customer.name).all()]
    
    if form.validate_on_submit():
        start_date = form.start_date.data
        end_date = form.end_date.data
        customer_id = form.customer.data
        include_fields = form.include_fields.data
        export_type = request.form.get('export_type', 'pdf')
        
        if not include_fields:
            include_fields = ['invoice_number', 'invoice_date', 'customer', 'subtotal', 'cgst', 'sgst', 'igst', 'total']
        
        invoices = Invoice.query
        
        if start_date:
            invoices = invoices.filter(Invoice.invoice_date >= start_date)
        if end_date:
            invoices = invoices.filter(Invoice.invoice_date <= end_date)
        if customer_id and customer_id != 0:
            invoices = invoices.filter(Invoice.customer_id == customer_id)
        
        invoices = invoices.order_by(Invoice.invoice_date.desc()).all()
        
        if export_type == 'excel':
            return generate_excel_report(invoices, include_fields, start_date, end_date)
        else:
            return generate_pdf_report(invoices, include_fields, start_date, end_date)
    
    return render_template('reports.html', form=form)

def generate_excel_report(invoices, include_fields, start_date, end_date):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Invoice Report"
        
        header_fill = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        headers = []
        for field in include_fields:
            if field == 'customer':
                headers.append('Customer')
            elif field == 'invoice_date':
                headers.append('Invoice Date')
            else:
                headers.append(field.replace('_', ' ').title())
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        for row, invoice in enumerate(invoices, 2):
            for col, field in enumerate(include_fields, 1):
                if field == 'customer':
                    ws.cell(row=row, column=col, value=invoice.customer.name)
                elif field == 'invoice_date':
                    ws.cell(row=row, column=col, value=invoice.invoice_date.strftime('%d-%m-%Y'))
                elif field == 'invoice_number':
                    ws.cell(row=row, column=col, value=invoice.invoice_number)
                elif field == 'subtotal':
                    ws.cell(row=row, column=col, value=float(invoice.subtotal))
                elif field == 'cgst':
                    ws.cell(row=row, column=col, value=float(invoice.cgst))
                elif field == 'sgst':
                    ws.cell(row=row, column=col, value=float(invoice.sgst))
                elif field == 'igst':
                    ws.cell(row=row, column=col, value=float(invoice.igst))
                elif field == 'total':
                    ws.cell(row=row, column=col, value=float(invoice.total))
                
                ws.cell(row=row, column=col).border = border
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 15
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = 'attachment; filename=invoice_report.xlsx'
        return response
        
    except ImportError:
        flash('Please install openpyxl to export Excel files: pip install openpyxl', 'warning')
        return redirect(url_for('reports'))

def generate_pdf_report(invoices, include_fields, start_date, end_date):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    story = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, spaceAfter=20, alignment=1)
    story.append(Paragraph("INVOICE REPORT", title_style))
    
    if start_date and end_date:
        story.append(Paragraph(f"Period: {start_date} to {end_date}", styles['Normal']))
    elif start_date:
        story.append(Paragraph(f"From: {start_date}", styles['Normal']))
    elif end_date:
        story.append(Paragraph(f"To: {end_date}", styles['Normal']))
    
    story.append(Spacer(1, 0.2*inch))
    
    headers = [field.replace('_', ' ').title() for field in include_fields]
    data = [headers]
    
    for invoice in invoices:
        row = []
        for field in include_fields:
            if field == 'customer':
                row.append(invoice.customer.name)
            elif field == 'invoice_date':
                row.append(invoice.invoice_date.strftime('%d-%m-%Y'))
            else:
                row.append(str(getattr(invoice, field)))
        data.append(row)
    
    if data and len(data) > 1:
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(table)
    
    doc.build(story)
    
    pdf = buffer.getvalue()
    buffer.close()
    
    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'attachment; filename=invoice_report.pdf'
    return response

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True)
